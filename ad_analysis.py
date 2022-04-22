import tkinter as tk # 图形界面库
from tkinter import filedialog # 图形界面库的文件对话框模块
from tkinter import messagebox # 图形界面库的弹窗模块
from win32com.client import Dispatch
from openpyxl import load_workbook
class demo(object):
    def __init__(self):
        
        window = tk.Tk()
        width = 300
        hight = 200
        screen_width = window.winfo_screenwidth()
        screen_hight = window.winfo_screenheight()
        info = "%dx%d+%d+%d" % (width,hight,(screen_width-width)/2,(screen_hight-hight)/2)
        window.geometry(info)
        window.title('Excel文件处理')
        # 创建一个按钮,按钮放在window窗口上
        # 按钮上的字是“选择文件”,按下后触发功能“cho_files”函数
        self.bo_cho = tk.Button(window, text='选择文件', command=self.cho_file)

        self.bo_deal = tk.Button(window, text='处理文件', command=self.deal_file)

        # 创建一个标签
        lable1 = tk.Label(window, text='选择要处理的文件，在点击处理')

        # 创建一个多行文本框,放在window窗口上
        self.Text_word = tk.Text(window)

        self.Text_word.insert('insert', '数据解释\n A：高点击,高CR\n B：高点击,低CR\n C：低点击,高CR\n D：低点击,低CR')
        # pack
        self.bo_cho.pack()
        self.bo_deal.pack()
        lable1.pack()
        self.Text_word.pack()

        self.file = None # 定义一个变量,来接收单个文件,例：'D:/product/文件1.xls'
        self.file_name = None # 用来接收文件名称,例：‘文件1.xls’
        self.file_path = None # 用来接收文件的路径,例：‘D:/product/’


        tk.mainloop()

    def cho_file(self):
        # 路径问题
        self.file = filedialog.askopenfilename()
        self.file_name = self.file.split('/')[-1]
        self.file_path = self.file.replace(self.file_name,"")
        messagebox.showinfo(title='文件已就绪！', message='单击处理文件。')
        print(self.file)


    def deal_file(self):
        #time_read1 = time.time()
        file2 = self.file_path +'\\'+'广告筛选1.5.1 - 轻量级.xlsm'
        wb = load_workbook(file2,keep_vba=True)
        wb2 = load_workbook(self.file,read_only=True)

        ws1= wb.sheetnames[1]
        ws2 = wb2.sheetnames[0]
        wa1 = wb[ws1]
        wa2 = wb2[ws2]

        if wa1['A1']:
            wa1.delete_cols(1,26)
        
        for row in wa2.iter_rows():
            row_list = []
            print(row)
            for cell in row:
                row_list.append(cell.value)
            wa1.append(row_list)


        wb.save(file2)
        
        
        
        # 弹出窗口,窗口标题为“成功”,窗口含有信息“成功生成文件。”
        messagebox.showinfo(title='成功', message='成功生成文件。')
        xlApp = Dispatch('Excel.Application')
        xlApp.Visible=1 # 显示excel界面
        xlBook = xlApp.Workbooks.Open(file2, ReadOnly = False) #打开对应的excel文件

if __name__==__name__:
    demo()